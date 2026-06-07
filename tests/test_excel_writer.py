from openpyxl import load_workbook

from dgts_crawler.excel_writer import (
    AuctionRow,
    HEADERS,
    SELECT_ORG_HEADERS,
    SELECT_ORG_RESULT_HEADERS,
    write_select_org_workbook,
    write_select_org_result_workbook,
    write_workbook,
)


def _row(sheet_name, notice_id, price):
    return AuctionRow(
        sheet_name=sheet_name,
        values=[
            1,
            "05/06/2026",
            "Lần 1",
            "Cơ quan A",
            "Tổ chức A",
            "Thông tin đấu giá",
            "Quyền sử dụng đất",
            "01",
            "Hà Nội",
            price,
            price // 10,
            "Ghi chú",
            "08/06/2026",
            "12/06/2026",
            "Đăng ký tại trụ sở",
            "10/06/2026",
            "11/06/2026",
            f"https://dgts.moj.gov.vn/thong-bao-cong-khai-viec-dau-gia/{notice_id}.html",
        ],
    )


def test_write_workbook_creates_single_sheet_with_generated_columns(tmp_path):
    output = tmp_path / "out.xlsx"
    rows = [
        AuctionRow(
            sheet_name="Ngân hàng",
            values=[
                1,
                "05/06/2026",
                "Lần 1",
                "Cơ quan A",
                "Tổ chức A",
                "Thông tin đấu giá",
                "Quyền sử dụng đất",
                "01",
                "Hà Nội",
                450000000,
                45000000,
                "Ghi chú",
                "08/06/2026",
                "12/06/2026",
                "Đăng ký tại trụ sở",
                "10/06/2026",
                "11/06/2026",
                "https://dgts.moj.gov.vn/thong-bao-cong-khai-viec-dau-gia/xe-1.html",
            ],
        ),
        AuctionRow(
            sheet_name="Thi hành án",
            values=[
                1,
                "05/06/2026",
                "Lần 1",
                "Ngân hàng A",
                "Tổ chức B",
                "Thông tin đấu giá",
                "Quyền sử dụng đất",
                "01",
                "Tây Ninh",
                579000000,
                57900000,
                "Ghi chú",
                "08/06/2026",
                "22/06/2026",
                "Đăng ký tại trụ sở",
                "10/06/2026",
                "11/06/2026",
                "https://dgts.moj.gov.vn/thong-bao-cong-khai-viec-dau-gia/dat-2.html",
            ],
        ),
    ]

    write_workbook(output, rows)

    wb = load_workbook(output)
    assert wb.sheetnames == ["Danh sách"]
    ws = wb["Danh sách"]
    assert [ws.cell(1, column).value for column in range(1, 19)] == HEADERS
    assert ws["A2"].value == 1
    assert ws["C2"].value == "Lần 1"
    assert ws["G2"].value == "Quyền sử dụng đất"
    assert ws["J2"].value == 450000000
    assert ws["K2"].value == 45000000
    assert ws["N3"].value == "22/06/2026"
    assert ws["R3"].value.endswith("dat-2.html")


def test_write_workbook_still_creates_headers_when_there_is_no_data(tmp_path):
    output = tmp_path / "empty.xlsx"

    write_workbook(output, [])

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert wb.sheetnames == ["Danh sách"]
    assert [ws.cell(1, column).value for column in range(1, 19)] == HEADERS
    assert ws.max_row == 1
    assert ws.auto_filter.ref == "A1:R1"


def test_write_workbook_applies_same_data_row_format_to_all_generated_rows(tmp_path):
    output = tmp_path / "out.xlsx"
    rows = [_row("Đất đai", i, 1000000 + i) for i in range(1, 9)]

    write_workbook(output, rows)

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    for column in range(1, 19):
        first = ws.cell(row=2, column=column)
        later = ws.cell(row=9, column=column)
        assert later._style == first._style
        assert later.number_format == first.number_format
        assert later.alignment.horizontal == first.alignment.horizontal
        assert later.alignment.vertical == first.alignment.vertical
        assert later.alignment.wrap_text == first.alignment.wrap_text


def test_write_workbook_formats_percent_deposit_cells(tmp_path):
    output = tmp_path / "out.xlsx"
    row = _row("Thanh lý", 1, 1000000)
    values = list(row.values)
    values[10] = "20%"

    write_workbook(output, [AuctionRow("Thanh lý", values)])

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert ws["K2"].value == 0.2
    assert ws["K2"].number_format == "0%"


def test_write_workbook_percent_deposit_does_not_leak_format_to_later_money_rows(tmp_path):
    output = tmp_path / "out.xlsx"
    percent_values = list(_row("Đất đai", 1, 1000000).values)
    percent_values[10] = "20%"
    money_values = list(_row("Đất đai", 2, 2000000).values)
    money_values[10] = 200000

    write_workbook(
        output,
        [AuctionRow("Đất đai", percent_values), AuctionRow("Đất đai", money_values)],
    )

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert ws["K2"].value == 0.2
    assert ws["K2"].number_format == "0%"
    assert ws["K3"].value == 200000
    assert ws["K3"].number_format != "0%"


def test_write_workbook_auto_expands_row_height_for_long_text(tmp_path):
    output = tmp_path / "out.xlsx"
    long_values = list(_row("Đất đai", 1, 1000000).values)
    long_values[6] = " ".join(["Thông báo đấu giá quyền sử dụng đất tại khu dân cư"] * 12)
    long_values[3] = " ".join(["Thi hành án dân sự tỉnh Cà Mau"] * 8)
    long_values[17] = "https://dgts.moj.gov.vn/" + "-".join(["duong-dan-chi-tiet"] * 16)

    write_workbook(output, [AuctionRow("Đất đai", long_values)])

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert ws.row_dimensions[2].height > 60
    assert ws.row_dimensions[2].height <= 220


def test_write_select_org_workbook_creates_thirteen_column_sheet(tmp_path):
    output = tmp_path / "select-org.xlsx"
    values = [
        0,
        "05/06/2026",
        "Tài sản A\nTài sản B",
        "Cơ quan A",
        "Địa chỉ A",
        "01\n02",
        "Tốt\nĐã qua sử dụng",
        4170200000,
        "08/06/2026",
        "10/06/2026",
        "Địa chỉ nhận HS",
        "Ông A - 0900000000",
        "https://dgts.moj.gov.vn/thong-bao-lua-chon-to-chuc-dau-gia/tai-san-a-1.html",
    ]

    write_select_org_workbook(output, [AuctionRow("", values)])

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert [ws.cell(1, column).value for column in range(1, 14)] == SELECT_ORG_HEADERS
    assert ws.max_column == 13
    assert ws["A2"].value == 1
    assert ws["C2"].value == "Tài sản A\nTài sản B"
    assert ws["H2"].value == 4170200000
    assert ws["H2"].number_format == "#,##0"
    assert ws["I2"].value == "08/06/2026"
    assert ws["J2"].value == "10/06/2026"
    assert ws.auto_filter.ref == "A1:M2"
    assert ws.row_dimensions[2].height > 22


def test_write_select_org_result_workbook_creates_twelve_column_sheet(tmp_path):
    output = tmp_path / "select-org-result.xlsx"
    values = [
        0,
        "05/06/2026",
        "Xe ô tô cứu thương BKS 14C 1248",
        "Bệnh viện Đa khoa khu vực Cẩm Phả",
        "Tổ 1, Khu 3, phường Cẩm Thịnh, Tỉnh Quảng Ninh",
        "01",
        "Đã qua sử dụng",
        19880000,
        "Trung tâm dịch vụ đấu giá tài sản Quảng Ninh",
        "Tầng 5, trụ sở Trung tâm Phục vụ hành chính công tỉnh Quảng Ninh",
        "0203.3508.336",
        "https://dgts.moj.gov.vn/thong-bao-ket-qua-lua-chon-to-chuc-dau-gia/xe-25941.html",
    ]

    write_select_org_result_workbook(output, [AuctionRow("", values)])

    wb = load_workbook(output)
    ws = wb["Danh sách"]
    assert [ws.cell(1, column).value for column in range(1, 13)] == SELECT_ORG_RESULT_HEADERS
    assert ws.max_column == 12
    assert ws["A2"].value == 1
    assert ws["H2"].value == 19880000
    assert ws["H2"].number_format == "#,##0"
    assert ws["I2"].value == "Trung tâm dịch vụ đấu giá tài sản Quảng Ninh"
    assert ws.auto_filter.ref == "A1:L2"
