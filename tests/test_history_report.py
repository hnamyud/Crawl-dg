from openpyxl import load_workbook

from dgts_crawler.history_store import HistoryEventRow
from dgts_crawler.history_report import HISTORY_REPORT_HEADERS, write_history_report


def test_write_history_report_creates_readable_excel_file(tmp_path):
    output = tmp_path / "history-report.xlsx"
    rows = [
        HistoryEventRow(
            created_at="2026-06-06T10:00:00",
            notice_kind="auction",
            event_type="CHANGED",
            notice_id="1",
            publish_date="05/06/2026",
            detail_url="https://dgts.moj.gov.vn/tin-1.html",
            changed_fields="asset_name",
            changed_details="asset_name:\nCũ: Tài sản A\nMới: Tài sản B",
            old_values="asset_name: Tài sản A",
            new_values="asset_name: Tài sản B",
            matched_notice_id="",
        )
    ]

    write_history_report(output, rows)

    wb = load_workbook(output)
    ws = wb["Lịch sử thay đổi"]
    assert [ws.cell(1, column).value for column in range(1, 12)] == HISTORY_REPORT_HEADERS
    assert ws["A2"].value == "2026-06-06T10:00:00"
    assert ws["C2"].value == "CHANGED"
    assert ws["G2"].value == "asset_name:\nCũ: Tài sản A\nMới: Tài sản B"
    assert ws["H2"].value == "asset_name: Tài sản A"
    assert ws["I2"].value == "asset_name: Tài sản B"
    assert ws["K2"].value == "https://dgts.moj.gov.vn/tin-1.html"
