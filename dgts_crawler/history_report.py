from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .history_store import HistoryEventRow


HISTORY_REPORT_HEADERS = [
    "Thời điểm phát hiện",
    "Loại tin",
    "Sự kiện",
    "ID tin",
    "Ngày đăng",
    "Trường thay đổi",
    "Chi tiết thay đổi",
    "Giá trị cũ",
    "Giá trị mới",
    "ID tin nghi trùng",
    "Đường dẫn chi tiết",
]
HISTORY_REPORT_WIDTHS = {
    1: 24,
    2: 20,
    3: 16,
    4: 16,
    5: 16,
    6: 34,
    7: 70,
    8: 60,
    9: 60,
    10: 18,
    11: 68,
}


def write_history_report(output_path: str | Path, rows: Iterable[HistoryEventRow]) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Lịch sử thay đổi"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column, width in HISTORY_REPORT_WIDTHS.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    _write_header(sheet)

    border = _thin_border()
    for row_index, row in enumerate(rows, start=2):
        values = [
            row.created_at,
            row.notice_kind,
            row.event_type,
            row.notice_id,
            row.publish_date,
            row.changed_fields,
            row.changed_details,
            row.old_values,
            row.new_values,
            row.matched_notice_id,
            row.detail_url,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 2, 3, 4, 5} else "left",
                vertical="center",
                wrap_text=True,
            )
    last_row = max(1, sheet.max_row)
    sheet.auto_filter.ref = f"A1:K{last_row}"
    wb.save(output)
    return output


def _write_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = _thin_border()
    for column, header in enumerate(HISTORY_REPORT_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border
    sheet.row_dimensions[1].height = 34


def _thin_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)
