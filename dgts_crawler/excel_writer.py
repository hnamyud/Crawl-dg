from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class AuctionRow:
    sheet_name: str
    values: list[Any]


SHEET_TITLE = "Danh sách"
HEADERS = [
    "STT",
    "Thời gian đăng",
    "Đăng lần",
    "Thông tin người có tài sản",
    "Thông tin đơn vị tổ chức hành nghề đấu giá",
    "Thông tin việc đấu giá",
    "Tên tài sản",
    "Số lượng",
    "Nơi có tài sản",
    "Giá khởi điểm",
    "Tiền đặt trước",
    "Ghi chú",
    "Thời gian bắt đầu đăng ký tham gia",
    "Thời gian kết thúc đăng ký tham gia",
    "Địa điểm, điều kiện, cách thức đăng ký",
    "Thời gian bắt đầu nộp tiền đặt trước",
    "Thời gian kết thúc nộp tiền đặt trước",
    "Url trang",
]
SELECT_ORG_HEADERS = [
    "STT",
    "Ngày đăng",
    "Tên tài sản",
    "Cơ quan có tài sản",
    "Địa chỉ",
    "Số lượng",
    "Chất lượng",
    "Giá khởi điểm",
    "Thời gian tiếp nhận HS",
    "Thời gian kết thúc HS",
    "Địa chỉ tiếp nhận HS",
    "Thông tin liên hệ",
    "Đường dẫn chi tiết",
]
SELECT_ORG_RESULT_HEADERS = [
    "STT",
    "Ngày đăng",
    "Tên tài sản",
    "Cơ quan có tài sản",
    "Địa chỉ",
    "Số lượng",
    "Chất lượng",
    "Giá khởi điểm",
    "Tổ chức được chọn",
    "Địa chỉ",
    "Thông tin liên hệ",
    "Đường dẫn chi tiết",
]
COLUMN_WIDTHS = {
    1: 8,
    2: 18,
    3: 12,
    4: 42,
    5: 42,
    6: 42,
    7: 60,
    8: 14,
    9: 44,
    10: 18,
    11: 18,
    12: 36,
    13: 20,
    14: 20,
    15: 44,
    16: 20,
    17: 20,
    18: 58,
}
SELECT_ORG_COLUMN_WIDTHS = {
    1: 8,
    2: 18,
    3: 62,
    4: 42,
    5: 42,
    6: 16,
    7: 46,
    8: 18,
    9: 20,
    10: 20,
    11: 44,
    12: 36,
    13: 58,
}
SELECT_ORG_RESULT_COLUMN_WIDTHS = {
    1: 8,
    2: 18,
    3: 62,
    4: 42,
    5: 42,
    6: 16,
    7: 46,
    8: 18,
    9: 44,
    10: 44,
    11: 34,
    12: 58,
}
AUTO_HEIGHT_MIN = 22
AUTO_HEIGHT_MAX = 220
AUTO_HEIGHT_LINE_POINTS = 18
AUTO_HEIGHT_COLUMNS = {
    4: 38,
    5: 38,
    6: 38,
    7: 54,
    9: 38,
    12: 32,
    15: 38,
    18: 44,
}
SELECT_ORG_AUTO_HEIGHT_COLUMNS = {
    3: 54,
    4: 38,
    5: 38,
    7: 40,
    11: 38,
    12: 32,
    13: 44,
}
SELECT_ORG_RESULT_AUTO_HEIGHT_COLUMNS = {
    3: 54,
    4: 38,
    5: 38,
    7: 40,
    9: 38,
    10: 38,
    11: 32,
    12: 44,
}


def write_workbook(output_path: str | Path, rows: Iterable[AuctionRow]) -> Path:
    return write_rows_workbook(
        output_path=output_path,
        sheet_title=SHEET_TITLE,
        headers=HEADERS,
        rows=(_row_values(index, row) for index, row in enumerate(rows, start=1)),
        column_widths=COLUMN_WIDTHS,
        numeric_columns={10, 11},
        auto_height_columns=AUTO_HEIGHT_COLUMNS,
    )


def write_select_org_workbook(output_path: str | Path, rows: Iterable[AuctionRow]) -> Path:
    return write_rows_workbook(
        output_path=output_path,
        sheet_title=SHEET_TITLE,
        headers=SELECT_ORG_HEADERS,
        rows=(_select_org_row_values(index, row) for index, row in enumerate(rows, start=1)),
        column_widths=SELECT_ORG_COLUMN_WIDTHS,
        numeric_columns={8},
        auto_height_columns=SELECT_ORG_AUTO_HEIGHT_COLUMNS,
    )


def write_select_org_result_workbook(output_path: str | Path, rows: Iterable[AuctionRow]) -> Path:
    return write_rows_workbook(
        output_path=output_path,
        sheet_title=SHEET_TITLE,
        headers=SELECT_ORG_RESULT_HEADERS,
        rows=(_select_org_result_row_values(index, row) for index, row in enumerate(rows, start=1)),
        column_widths=SELECT_ORG_RESULT_COLUMN_WIDTHS,
        numeric_columns={8},
        auto_height_columns=SELECT_ORG_RESULT_AUTO_HEIGHT_COLUMNS,
    )


def write_rows_workbook(
    output_path: str | Path,
    sheet_title: str,
    headers: list[str],
    rows: Iterable[list[Any]],
    column_widths: dict[int, int],
    numeric_columns: set[int],
    auto_height_columns: dict[int, int],
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    _setup_sheet(sheet, headers, column_widths)
    _write_rows(sheet, list(rows), numeric_columns, auto_height_columns)
    wb.save(output)
    return output


def _setup_sheet(sheet: Any, headers: list[str], column_widths: dict[int, int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    _write_header(sheet, headers)


def _write_header(sheet: Any, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = _thin_border()

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border
    sheet.row_dimensions[1].height = 34


def _write_rows(
    sheet: Any,
    rows: list[list[Any]],
    numeric_columns: set[int],
    auto_height_columns: dict[int, int],
) -> None:
    border = _thin_border()
    for index, values in enumerate(rows, start=1):
        excel_row = index + 1
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=excel_row, column=column)
            _write_cell_value(cell, value)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = _data_alignment(column)
            if column in numeric_columns and isinstance(cell.value, (int, float)) and cell.number_format != "0%":
                cell.number_format = '#,##0'
        _auto_fit_data_row_height(sheet, excel_row, auto_height_columns)
    last_row = max(1, len(rows) + 1)
    last_column = get_column_letter(max(1, len(sheet[1])))
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"


def _row_values(index: int, row: AuctionRow) -> list[Any]:
    source = list(row.values)
    while len(source) < 18:
        source.append("")
    values = source[:18]
    values[0] = index
    return values


def _select_org_row_values(index: int, row: AuctionRow) -> list[Any]:
    source = list(row.values)
    while len(source) < 13:
        source.append("")
    values = source[:13]
    values[0] = index
    return values


def _select_org_result_row_values(index: int, row: AuctionRow) -> list[Any]:
    source = list(row.values)
    while len(source) < 12:
        source.append("")
    values = source[:12]
    values[0] = index
    return values


def _write_cell_value(cell: Any, value: Any) -> None:
    if isinstance(value, str) and value.strip().endswith("%"):
        numeric_text = value.strip()[:-1].replace(",", ".")
        try:
            cell.value = float(numeric_text) / 100
            cell.number_format = "0%"
            return
        except ValueError:
            pass
    cell.value = value


def _data_alignment(column: int) -> Alignment:
    horizontal = "center" if column in {1, 2, 3, 8, 13, 14, 16, 17} else "left"
    if column in {10, 11}:
        horizontal = "right"
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _thin_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def _auto_fit_data_row_height(sheet: Any, row_number: int, auto_height_columns: dict[int, int]) -> None:
    max_lines = 1
    for column, chars_per_line in auto_height_columns.items():
        value = sheet.cell(row=row_number, column=column).value
        if value in (None, ""):
            continue
        text = str(value)
        explicit_lines = text.count("\n") + 1
        wrapped_lines = max(1, (len(text) + chars_per_line - 1) // chars_per_line)
        max_lines = max(max_lines, explicit_lines, wrapped_lines)
    height = max(AUTO_HEIGHT_MIN, min(AUTO_HEIGHT_MAX, max_lines * AUTO_HEIGHT_LINE_POINTS))
    sheet.row_dimensions[row_number].height = height
